from tkinter import *
import tkinter
from tkinter import messagebox
import openpyxl
import os
from PIL import Image, ImageTk
from tkinter import ttk

#ttk = themed tkinter [collection of themed widgets for advancements]

# pyinstaller
# from main import markAttendance

dash = Tk()
dash.geometry("1440x1024")
dash.resizable(False,False)
dash.title("Dashboard")

# Navigation Bar
Label(text="Face Detection Attendance System",bg="#05536D", fg="#D6E0E3", font="Inter 35", padx=25, pady=17, borderwidth=0.8, relief=GROOVE).pack(fill=X)



# Registeration Window
def reg_window():

    dash.destroy()

    def validate():
        ernum = ernum_entry.get()
        email = email_entry.get()
        department = department_entry.get()
        branch = branch_entry.get()
        # VALIDATION 
        if ernum:
            print("ER number: ", ernum)
            if email:
                print("Email: ", email)
                if department:
                    print("Department: ", department)
                    if branch:
                        print("Branch: ", branch)
                        
                        path= "E:/VSCode/FDAS/registeration.xlsx"

                        if not os.path.exists(path):
                            excel = openpyxl.Workbook()
                            regsheet = excel.active
                            heading = ["ER Num", "E-mail", "Department", "Branch"]
                            regsheet.append(heading)
                            excel.save(path)
                        
                        excel = openpyxl.load_workbook(path)
                        regsheet= excel.active
                        regsheet.append([ernum, email, department, branch])
                        excel.save(path)


                        tkinter.messagebox.showwarning(title="Success", message="User Registered")

                    else:
                        tkinter.messagebox.showwarning(title="Warning", message="Enter Branch")
                else:
                    tkinter.messagebox.showwarning(title="Warning", message="Enter Department")
            else:
                tkinter.messagebox.showwarning(title="Warning", message="Enter Email")
        else:
            tkinter.messagebox.showwarning(title="Warning", message="Enter ER number")

    


    regbox = tkinter.Tk()
    regbox.title("registeration form")


    Label(regbox,text="Registration Form",bg="#05536D", fg="#D6E0E3", font="Inter 35", padx=25, pady=17, borderwidth=0.8, relief=GROOVE).pack(fill=X)



    # where everything will be inserted
    bluebox = tkinter.Frame(regbox)
    bluebox.pack()
    # bluebox.place(x=761,y=120,width=460,height=540)

    # All entitie [ernum, email, department, branch] information storing.
    info = tkinter.LabelFrame(bluebox, text= "Information")
    info.grid(row= 0, column=0, padx=20, pady=20) 


    ernum_label = Label(info, text="Enrollment no:", font=("Inter",12,"bold"))      #fg='#D6E0E3', bg='#05536D' Colors
    ernum_label.grid(row=0, column=0)
    ernum_entry = tkinter.Entry(info)
    ernum_entry.grid(row=0, column=1)

    email_label = Label(info, text="Email Address:", font=("Inter",12,"bold"))
    email_label.grid(row=1, column=0)
    email_entry = tkinter.Entry(info)
    email_entry.grid(row=1, column=1)

    department_label = Label(info, text="Department:", font=("Inter",12,"bold"))
    department_label.grid(row=2, column=0)
    department_entry = tkinter.Entry(info)
    department_entry.grid(row=2, column=1)

    branch_label = Label(info, text="Branch:", font=("Inter",12,"bold"))
    branch_label.grid(row=3, column=0)
    branch_entry = tkinter.Entry(info)
    branch_entry.grid(row=3, column=1)


    button = tkinter.Button(bluebox, text="Register", command= validate)
    button.grid(row=4, column=0, padx=20, pady=10)

    button = tkinter.Button(bluebox, text="Back", command= exit)            #isme exit ki jagah par ek back button banana hai jo dashboard pe le jaaye
    button.grid(row=5, column=0, padx=20, pady=10)


# ATTENDANCE WINDOW MA IMAGES LEVANA CHE...
# NIZA KARI APJE E TU



def att_window():
#     # Destroys Dashboard Window
    dash.destroy()
    # reg.destroy()

    att = Tk()
    att.title("Attendance Window")
    Label(att,text="Attendance Window", bg="#05536D", fg="#D6E0E3", font="Inter 35", padx=25, pady=17, borderwidth=0.8, relief=GROOVE).pack(fill=X)
    att.geometry("700x700")

    # Take Attendance Window
    attlog = Image.open("attendanceImg.png")
    attlogres = attlog.resize((360,380))
    tkatt = ImageTk.PhotoImage(attlogres)
    # Button(att,text='Take Attendance', image=tkatt, =0, command=markAttendance).place(x=200,y=200)
    Button(att,text='Take Attendance', image=tkatt,width=200, height=200, border=0, command=Att).place(x=200,y=200)

    # View Attendance Window
    viatt = Image.open("regImg.png")
    viewlogres = viatt.resize((360,380))
    tkatt = ImageTk.PhotoImage(viewlogres)
    Button(att,text='View Attendance', image=tkatt, border=0, command=viewdata).place(x=200,y=200)


    # Buttons  
    takeatt = LabelFrame(att,bd=2,bg="#05536D",font=("Inter",30,"bold"))
    takeatt.place(x=761,y=120,width=460,height=540)
    Button(takeatt, text="Take Attendance", command=exit,width=15,fg="#FFFFFF", bg="#F85900", font=("Inter",12)).place(x=170,y=420)

    takeatt = LabelFrame(att,bd=2,bg="#05536D",font=("Inter",30,"bold"))
    takeatt.place(x=1400,y=120,width=460,height=540)
    Button(takeatt, text="View Attendance", command=exit,width=15,fg="#FFFFFF", bg="#F85900", font=("Inter",12)).place(x=170,y=420)

    # markAttendance


# View Data from excel
def viewdata():
    path = "Attendance.xlsx"
    excel = openpyxl.load_workbook(path)
    viewsheet = excel.active
    values = list(viewsheet.values)
    columns = values[0]
    treeview = ttk.Treeview(dash, columns=columns, show="headings")
    for valrows in values[1:]:
        treeview.insert("", tkinter.END, values=valrows)


# Attendance Button in Dashboard
attendance_logo = Image.open("attendanceImg.png")
attendance_logo_resized = attendance_logo.resize((360,380))
attendance_image = ImageTk.PhotoImage(attendance_logo_resized)
Button(dash,text='Register', image=attendance_image, border=0, command=att_window).place(x=200,y=200)

# Registration Button in Dashboard
reg_logo = Image.open("regImg.png")
reg_logo_resized = reg_logo.resize((360,380))
reg_image = ImageTk.PhotoImage(reg_logo_resized)
Button(dash,text='Register', image=reg_image, border=0, command=reg_window).place(x=700,y=200)



dash.mainloop()